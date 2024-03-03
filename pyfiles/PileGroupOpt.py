import numpy as np
import openpyxl
import time
import math

import calfem.core as cfc
import matplotlib.pyplot as plt

class PileOptModel:

    def defineSettings(self,xvec,yvec,npiles,nvert,singdir,plen,incl,path,nrVal,Nmax,Nmin):

        # Materialegenskaper
        E                       = 2.0e11
        A1                      = 1.2193e-2
        
        A                       = 1000000.0
        G                       = 1000000.0
        Iy                      = 1000000.0
        Iz                      = 1000000.0
        Kv                      = 1000000.0
        
        self.ep1                = [E, G, A, Iy, Iz, Kv]
        self.ep2                = [E, A1]

        # Lastindata
        self.path               = path
        self.nrVal              = nrVal

        # Indata en kvadrant
        self.x1vec_q_s          = xvec
        self.y1vec_q_s          = yvec

        self.npiles_q           = npiles
        self.npiles             = npiles*4

        self.incl               = incl
        self.nVertPiles         = nvert
        self.nrSingDirPiles     = singdir
        self.pLen               = plen

        self.Nmaxval            = Nmax
        self.Nminval            = Nmin

    def genElem(self):

        # Elementdata, beam elements
        self.Coord1             = np.array(np.zeros((self.npiles+1, 3)))
        self.Dof1               = np.array(np.zeros((self.npiles+1, 6)))
        self.Edof1              = np.array(np.zeros((self.npiles, 12)))

        # Elementdata, bar elements
        self.Coord2             = np.array(np.zeros((2*self.npiles, 3)))
        self.Dof2               = np.array(np.zeros((2*self.npiles, 3)))
        self.Edof2              = np.array(np.zeros((self.npiles, 6)))
        
        dpap                    = 0.000001
        self.Coord1[0, 2]       = dpap

        for i in range(self.npiles):
            self.Coord1[i+1, 0] = self.x1vec[i]
            self.Coord1[i+1, 1] = self.y1vec[i]

        for i in range(self.npiles+1):
            self.Dof1[i, :] = [6*i+1, 6*i+2, 6*i+3, 6*i+4, 6*i+5, 6*i+6]
        
        for i in range(self.npiles):
            self.Edof1[i, :] = [1, 2, 3, 4, 5, 6, 6*i+7, 6*i+8, 6*i+9, 6*i+10, 6*i+11, 6*i+12]
        
        [self.Ex1, self.Ey1, self.Ez1] = cfc.coordxtr(self.Edof1, self.Coord1, self.Dof1, 2)
        
        for i in range(self.npiles):
            if self.incl[i] == 0:
                self.x2vec[i] = self.x1vec[i]
                self.y2vec[i] = self.y1vec[i]
            else:
                self.x2vec[i] = self.x1vec[i] + np.cos(np.radians(self.bearing[i]))*self.lvec[i]/self.incl[i]
                self.y2vec[i] = self.y1vec[i] + np.sin(np.radians(self.bearing[i]))*self.lvec[i]/self.incl[i]
        
        for i in range(self.npiles):
            self.Coord2[i, 0] = self.x1vec[i]
            self.Coord2[i, 1] = self.y1vec[i]
        
            self.Coord2[i+self.npiles, 0] = self.x2vec[i]
            self.Coord2[i+self.npiles, 1] = self.y2vec[i]
            self.Coord2[i+self.npiles, 2] = -self.lvec[i]
        
        for i in range(2*self.npiles):
            self.Dof2[i, :] = [6*i+7, 6*i+8, 6*i+9]
        
        for i in range(self.npiles):
            self.Edof2[i, :] = [6*i+7, 6*i+8, 6*i+9, 6*i+(6*self.npiles)+7, 6*i+(6*self.npiles)+8, 6*i+(6*self.npiles)+9]
        
        [self.Ex2, self.Ey2, self.Ez2] = cfc.coordxtr(self.Edof2, self.Coord2, self.Dof2, 2)

        self.nDofs = int(np.max(self.Edof2))
        

    def assembElem(self):    

        self.K          = np.matrix(np.zeros((self.nDofs, self.nDofs)))
        self.Edof1      = self.Edof1.astype(int)
        
        for i in range(self.npiles):
        
            p1 = np.array([0, 0, 0])
            p2 = np.array([self.Ex1[i, :][0], self.Ey1[i, :][0], self.Ez1[i, :][0]])
            p3 = np.array([self.Ex1[i, :][1], self.Ey1[i, :][1], self.Ez1[i, :][1]])

            v1 = np.subtract(p2, p1)
            v2 = np.subtract(p3, p1)
            eo = np.cross(v1, v2)
        
            Ke = cfc.beam3e(self.Ex1[i, :], self.Ey1[i, :], self.Ez1[i, :], eo, self.ep1)
            self.K = cfc.assem(self.Edof1[i, :], self.K, Ke)
        
        self.Edof2 = self.Edof2.astype(int)
        
        for i in range(self.npiles):

            p1 = np.array([0, 0, 0])
            p2 = np.array([self.Ex2[i, :][0], self.Ey2[i, :][0], self.Ez2[i, :][0]])
            p3 = np.array([self.Ex2[i, :][1], self.Ey2[i, :][1], self.Ez2[i, :][1]])
        
            Ke = cfc.bar3e(self.Ex2[i, :], self.Ey2[i, :], self.Ez2[i, :], self.ep2)
            self.K = cfc.assem(self.Edof2[i, :], self.K, Ke)

    def readLoadCases(self):
        print("- Reading loadcases...")
        # Läser av excelark och genererar matris med samtliga lastfall, kaliberar manuellt
        wb = openpyxl.load_workbook(self.path, data_only=True)
        sheet = wb.active
        self.lc = np.array(np.zeros((self.nrVal, 7)))

        for i in range(self.nrVal):
            FX = sheet.cell(row=4+i, column=3).value
            FY = sheet.cell(row=4+i, column=4).value
            FZ = sheet.cell(row=4+i, column=5).value
            MX = sheet.cell(row=4+i, column=6).value
            MY = sheet.cell(row=4+i, column=7).value
            MZ = sheet.cell(row=4+i, column=8).value
            self.lc[i, :] = [i+1, FX, FY, FZ, MX, MY, MZ]

    def generateLoads(self,nr):

        f = np.array(np.zeros((self.nDofs, 1)))

        f[0] = self.lc[nr,1]*1000
        f[1] = self.lc[nr,2]*1000
        f[2] = self.lc[nr,3]*1000
        f[3] = self.lc[nr,4]*1000
        f[4] = self.lc[nr,5]*1000
        f[5] = self.lc[nr,6]*1000

        return f


    def analyseLoadcases(self,f):
        [a, r] = cfc.solveq(self.K, f, self.bc)

        return a, r


    def analyseResults(self,a):

        Nvek = np.array(np.zeros((self.npiles)))

        for i in range(self.npiles):
            Ed = cfc.extractEldisp(self.Edof2[i, :], a)
            N  = cfc.bar3s(self.Ex2[i, :], self.Ey2[i, :], self.Ez2[i, :], self.ep2, Ed)

            Nvek[i] = N[0][0]

        return Nvek

    def plotPileGroup(self):

        t0 = time.time()
        #plt.close()
        self.fig = plt.figure(dpi=100) # figsize=(20,40)
        #self.fig.clear()
        elapsed = time.time() - t0
        print(elapsed)

        ax = self.fig.subplots()

        xmax = max(self.x1vec_q_s)*1.15
        ymax = max(self.y1vec_q_s)*1.15

        ax.set_xlim(-xmax,xmax)
        ax.set_ylim(-ymax,ymax)

        Ex2normVek = self.Ex2
        Ey2normVek = self.Ey2

        fak = 10

        for i in range(self.npiles):
            ax.plot(self.Ex2[i][0], self.Ey2[i][0], 'o', mfc='none', color='black',markersize=8)
            ax.set_aspect('equal')
            
            Ex2norm = self.Ex2[i,0] - self.Ex2[i,1]
            Ey2norm = self.Ey2[i,0] - self.Ey2[i,1]

            Ex2normVek[i,1] = self.Ex2[i,0] - Ex2norm/fak
            Ey2normVek[i,1] = self.Ey2[i,0] - Ey2norm/fak

            ax.plot(Ex2normVek[i,:], Ey2normVek[i,:], color='black')

        #plt.show()

    def pileExpand(self,nr):
        # Expanding pile data from a single quadrant

        a = [1, -1, -1, 1]
        b = [1, 1, -1, -1]
        c = [0, 90, 180, 270]

        # Kvadrantdata
        self.bearing_q      = np.array(np.zeros((self.npiles_q)))
        self.x1vec_q        = np.array(np.zeros((self.npiles_q)))
        self.y1vec_q        = np.array(np.zeros((self.npiles_q)))
        self.x2vec_q        = np.array(np.zeros((self.npiles_q)))
        self.y2vec_q        = np.array(np.zeros((self.npiles_q)))
        self.z1vec_q        = np.array(np.zeros((self.npiles_q)))
        self.incl_q         = np.array(np.zeros((self.npiles_q)))
        self.lvec_q         = np.ones((self.npiles_q))*self.pLen

        # Full data 
        self.bearing        = np.array(np.zeros((self.npiles_q*4)))
        self.x1vec          = np.array(np.zeros((self.npiles_q*4)))
        self.y1vec          = np.array(np.zeros((self.npiles_q*4)))
        self.z1vec          = np.array(np.zeros((self.npiles_q*4)))
        self.x2vec          = np.array(np.zeros((self.npiles_q*4)))
        self.y2vec          = np.array(np.zeros((self.npiles_q*4)))
        self.incl           = np.array(np.zeros((self.npiles_q*4)))
        self.lvec           = np.array(np.zeros((self.npiles_q*4)))

        # Set current pile config from arr-vec
        self.bearing_q      = self.bearing_arr[nr]
        self.x1vec_q        = self.xvec_arr[nr]
        self.y1vec_q        = self.yvec_arr[nr]
        self.incl_q         = self.incl_arr[nr]

        for i in range(self.npiles_q):
            if self.incl_q[i] == 0:
                self.x2vec_q[i] = self.x1vec_q[i]
                self.y2vec_q[i] = self.y1vec_q[i]
            else:
                self.x2vec_q[i] = self.x1vec_q[i] + np.cos(np.radians(self.bearing_q[i]))*self.pLen/self.incl_q[i]
                self.y2vec_q[i] = self.y1vec_q[i] + np.sin(np.radians(self.bearing_q[i]))*self.pLen/self.incl_q[i]

        iter = 0
        for i in range(4):
            for j in range(self.npiles_q):

                self.x1vec[iter]    = a[i]*self.x1vec_q[j]
                self.y1vec[iter]    = b[i]*self.y1vec_q[j]
                self.x2vec[iter]    = a[i]*self.x2vec_q[j]
                self.y2vec[iter]    = b[i]*self.y2vec_q[j]
                self.z1vec[iter]    = self.z1vec_q[j]
                self.incl[iter]     = self.incl_q[j]
                self.lvec[iter]     = self.lvec_q[j]

                if i > 0:
                    if self.bearing[iter-self.npiles_q] == c[i]:
                        self.bearing[iter] = self.bearing[iter-self.npiles_q]
                    else:
                        self.bearing[iter] = self.bearing[iter-self.npiles_q] + 180
                else:
                    self.bearing[iter] = self.bearing_q[j]

                iter = iter + 1
    
    def pileInfluenceRun(self):
        
        print("- Running influence analysis...")

        self.configStore = []
        self.Nmaxstore = []
        self.Nminstore = []

        for config in range(self.nrConfigs):
            self.pileExpand(config)
            self.genElem()  
            
            self.f                  = np.matrix(np.zeros((self.nDofs, 1)))
            self.bc                 = np.arange((6*self.npiles)+7, self.nDofs+1)
            self.K                  = np.matrix(np.zeros((self.nDofs, self.nDofs)))

            self.assembElem()

            Ninfl = np.array(np.zeros((self.npiles,6)))

            for i in range(6):
                f = np.matrix(np.zeros((self.nDofs, 1)))
                f[i] = 1000
                a, r = self.analyseLoadcases(f)
                Ninfl[:,i] = self.analyseResults(a)

            Nvek = np.array(np.zeros((self.npiles,self.nrVal)))

            for i in range(self.nrVal):
                for j in range(self.npiles):
                    Nvek[j,i] = np.matmul(Ninfl[j,:],self.lc[i,1:])

            Nmax = round(0.001*np.max(Nvek))
            Nmin = round(0.001*np.min(Nvek))

            if Nmax < self.Nmaxval and Nmin > self.Nminval:
                print("Konfiguration " + str(config) + ": " + str(self.bearing_q) + " | " + str(self.incl_q) + " | " + str(self.set_arr[config]) + " | " + str(Nmax) + ", " + str(Nmin))
                self.configStore.append(config)
                self.Nmaxstore.append(Nmax)
                self.Nminstore.append(Nmin)


    def pileSolver(self,config):

        print("- Running single config analysis...")

        self.pileExpand(config)
        self.genElem()

        self.f                  = np.matrix(np.zeros((self.nDofs, 1)))
        self.bc                 = np.arange((6*self.npiles)+7, self.nDofs+1)
        self.K                  = np.matrix(np.zeros((self.nDofs, self.nDofs)))

        self.assembElem()

        Nmat = np.array(np.zeros((self.npiles,self.nrVal)))

        for i in range(self.nrVal):
            f = self.generateLoads(i)
            a, r = self.analyseLoadcases(f)
            Nmat[:,i] = self.analyseResults(a)

        Nmax = round(0.001*np.max(Nmat))
        Nmin = round(0.001*np.min(Nmat))

        print("Konfiguration " + str(config) + ": " + str(Nmax) + ", " + str(Nmin))

        self.plotPileGroup()



    def checkCollision(self,bearing_try,incl_try,x1vec_q,y1vec_q,dirvec):

        prec = 1
        step = max((x1vec_q[1] - x1vec_q[0]),(y1vec_q[1] - y1vec_q[0]))

        for dir in dirvec:

            z = dir*step*self.incl*0.5

            xy2vec_q = set()

            for i in range(self.npiles_q):
                if incl_try[i] == 0:
                    x2vec_q = x1vec_q[i]
                    y2vec_q = y1vec_q[i]
                else:
                    x2vec_q = x1vec_q[i] + np.cos(np.radians(bearing_try[i]))*z/incl_try[i]
                    y2vec_q = y1vec_q[i] + np.sin(np.radians(bearing_try[i]))*z/incl_try[i]

                if x2vec_q <= 0 or y2vec_q <= 0:
                    return True

                xy2vec_q.add((round(x2vec_q, prec), round(y2vec_q, prec)))

                if len(xy2vec_q) != i+1:
                    return True
            
        return False
    
    def genPileConfigs(self):

        print("- Finding and filtering possible pile configurations...")

        self.bearing_arr    = []
        self.incl_arr       = []
        self.xvec_arr       = []
        self.yvec_arr       = []
        self.set_arr        = []

        t0 = time.time()

        nInclPiles = self.npiles_q-self.nVertPiles

        dirvec = [1,2,3,4,-1,-2]
        #dirvec = [1,2,3,4,-1,-2,-3]

        npos = len(self.x1vec_q_s)

        # Pileset matrix
        setnr = int(math.factorial(npos) / (math.factorial(npos-self.npiles_q)*math.factorial(self.npiles_q)))
        pilesetmat = np.array(np.zeros((setnr,npos)))
        
        iter = 0
        for i in range(1 << npos):
            set_bin = format(i, '0' + str(npos) + 'b')
            set_try = [int(x) for x in [*set_bin]]
            if sum(set_try) == self.npiles_q:
                pilesetmat[iter,:] = set_try
                iter = iter + 1 

        # Pilexvec and Pileyvec
        pilexvec = []; pileyvec = []; setvec = []
        for i in range(setnr):
            pileset = pilesetmat[i]
            xvec = []; yvec = []
            iter = 0
            for j in range(len(pileset)):
                if pileset[j] != 0:
                    xvec.append(self.x1vec_q_s[j])
                    yvec.append(self.y1vec_q_s[j])

                    iter = iter + 1 
            pilexvec.append(xvec)
            pileyvec.append(yvec)
            setvec.append(pileset)

        # Pileincl matrix
        pileinclmat = []
        for i in range(1 << self.npiles_q):
            incl_bin = format(i, '0' + str(self.npiles_q) + 'b')
            incl_form = [int(x) for x in [*incl_bin]]
            if sum(incl_form) == nInclPiles:
                incl_try = [x*self.incl for x in incl_form]
                pileinclmat.append(incl_try)

        # Piledir matrix
        piledirmat = []
        for i in range(1 << nInclPiles):
            bearing_bin = format(i, '0' + str(nInclPiles) + 'b')
            bearing_form = [int(x) for x in [*bearing_bin]]
            npilesDir = sum(bearing_form)
            if npilesDir >= self.nrSingDirPiles and npilesDir <= self.npiles_q-self.nrSingDirPiles:
                bearing_try_temp = [x*90 for x in bearing_form]
                piledirmat.append(bearing_try_temp)

        # Generating possible configurations
        n = 0
        for i in range(setnr):
            x1vec_q = pilexvec[i]
            y1vec_q = pileyvec[i]
            set_try = setvec[i]
            for incl_try in pileinclmat:
                for bearing_try_temp in piledirmat:
                    bearing_try = np.zeros((self.npiles_q))
                    step = 0
                    n = n + 1
                    for j in range(self.npiles_q):
                        if incl_try[j] != 0:
                            bearing_try[j] = bearing_try_temp[step]
                            step = step + 1

                    # Checking for collisions to filter out only suitable configurations
                    if self.checkCollision(bearing_try,incl_try,x1vec_q,y1vec_q,dirvec) == False:
                        self.bearing_arr.append(bearing_try)
                        self.incl_arr.append(incl_try)
                        self.xvec_arr.append(x1vec_q)
                        self.yvec_arr.append(y1vec_q)
                        self.set_arr.append(set_try)

        self.nrConfigs = len(self.bearing_arr)

        elapsed = time.time() - t0
        print("- Number of possible configurations: " + str(self.nrConfigs) + " of " + str(n) + " | " + str(round(elapsed, 2)) + "s")

    

#self = PileOptModel()